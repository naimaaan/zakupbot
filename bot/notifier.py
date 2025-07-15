# notifier.py

import os
import re
import json
import requests
from datetime import datetime
from data_sources.test_api_fetch import fetch_procurement_plans
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from copy import copy

# Маппинг
DURATION_TYPE_MAP = {
    "ANNUAL": "Годовой план",
    "LONG_TIME": "Долгосрочный план",
}

TYPE_MAP = {
    "PREBASIC": "Предварительный",
    "BASIC": "Основной",
    "REVIEWED": "Уточнённый"
}

TRU_CODES = ["801019.000.000010"]

DOWNLOAD_DIR = "storage/downloads"
NOTIFIED_FILE = "storage/notified_uids.json"
os.makedirs(DOWNLOAD_DIR, exist_ok=True)
os.makedirs("storage", exist_ok=True)

def download_excel_file(uid: str) -> str | None:
    url = f"https://zakup.sk.kz/eprocfilestorage/open-api/files/download/{uid}"
    local_path = os.path.join(DOWNLOAD_DIR, f"{uid}.xlsx")

    try:
        response = requests.get(url)
        response.raise_for_status()
        with open(local_path, "wb") as f:
            f.write(response.content)
        return local_path
    except Exception as e:
        print(f"❌ Ошибка при скачивании Excel-файла {uid}: {e}")
        return None

def filter_excel_by_tru(filepath: str, tru_codes: list[str], save_file: bool = True) -> str | bool | None:
    try:
        wb = load_workbook(filepath)
        ws = wb.active

        output_path = filepath.replace(".xlsx", "_filtered.xlsx")
        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.title = "Filtered"

        for row_idx in range(1, 11):
            for col_idx, cell in enumerate(ws[row_idx], 1):
                new_cell = new_ws.cell(row=row_idx, column=col_idx, value=cell.value)
                new_cell.font = copy(cell.font)
                new_cell.alignment = copy(cell.alignment)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)

        for merged_range in ws.merged_cells.ranges:
            if merged_range.max_row <= 10:
                new_ws.merge_cells(str(merged_range))

        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            if ws.column_dimensions[letter].width:
                new_ws.column_dimensions[letter].width = ws.column_dimensions[letter].width

        insert_row = 11
        for row in ws.iter_rows(min_row=11):
            row_values = [str(cell.value) for cell in row]
            if any(tru in val for tru in tru_codes for val in row_values):
                for col_idx, cell in enumerate(row, 1):
                    new_cell = new_ws.cell(row=insert_row, column=col_idx, value=cell.value)
                    new_cell.font = copy(cell.font)
                    new_cell.alignment = copy(cell.alignment)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                insert_row += 1

        if insert_row == 11:
            return None

        if save_file:
            new_wb.save(output_path)
            return output_path

        return True

    except Exception as e:
        print(f"❌ Ошибка при фильтрации: {e}")
        return None

def get_procurement_summary(tru_codes: list[str]) -> list[dict]:
    plans = fetch_procurement_plans()
    messages = []

    for plan in plans:
        uid = plan.get("excelFileUid")
        if not uid:
            continue

        file_path = download_excel_file(uid)
        filtered_path = filter_excel_by_tru(file_path, tru_codes)

        # Если не найдено нужного ТРУ — пропускаем
        if not filtered_path:
            os.remove(file_path)
            continue

        # Данные для сообщения
        raw_date = plan.get("approveDate")
        date_time = datetime.fromtimestamp(raw_date / 1000).strftime("%Y-%m-%d %H:%M") if raw_date else "—"
        customer = plan.get("customerName", "—")
        customer_bin = plan.get("customerIdentifier", "—")
        year = plan.get("year", "—")
        duration_type = DURATION_TYPE_MAP.get(plan.get("planDurationType", "—"), "—")
        plan_type = TYPE_MAP.get(plan.get("planType", "—"), "—")

        message = (
            f"🏢  {customer}\n"
            f"🆔  БИН: {customer_bin}\n"
            f"📅  {date_time}\n"
            f"📋  {duration_type} | {plan_type} | {year}\n"
            f"🛡️  ТРУ: Услуги по обеспечению информационной безопасности.\n"
            f"🌐  Источник: zakup.sk.kz\n"
        )

        messages.append({"text": message, "uid": uid})

        os.remove(file_path)
        if filtered_path and os.path.exists(filtered_path):
            os.remove(filtered_path)

    return messages

def load_notified_uids() -> set:
    if not os.path.exists(NOTIFIED_FILE):
        return set()
    with open(NOTIFIED_FILE, "r", encoding="utf-8") as f:
        return set(json.load(f))

def save_notified_uids(uids: set):
    with open(NOTIFIED_FILE, "w", encoding="utf-8") as f:
        json.dump(list(uids), f, indent=2)


def extract_tru_rows(filepath: str) -> list[str]:
    try:
        wb = load_workbook(filepath)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(min_row=11, values_only=True):
            row_values = [str(cell) if cell is not None else "" for cell in row]
            row_text = " | ".join(row_values).strip()
            if any(tru in row_text for tru in TRU_CODES):
                rows.append(row_text)
        return rows
    except Exception as e:
        print(f"❌ Error extracting TРУ rows: {e}")
        return []
TRU_STORAGE_FILE = "storage/tru_rows.json"

def load_tru_history() -> dict[str, list[str]]:
    if not os.path.exists(TRU_STORAGE_FILE):
        return {}
    with open(TRU_STORAGE_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

def save_tru_history(data: dict[str, list[str]]):
    with open(TRU_STORAGE_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
